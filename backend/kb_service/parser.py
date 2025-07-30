import logging
import io
import rarfile
import zipfile
import magic
from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

def parse_document(file_name: str, file_content: bytes, mime_type: str) -> str:
    logger.info(f"--- PARSER START: Parsing '{file_name}' with MIME type: {mime_type} ---")
    
    try:
        SUPPORTED_TYPES = [
            'text/plain', 
            'application/pdf', 
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document', 
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
            'text/csv', 
            'application/zip', 
            'application/x-rar-compressed', 
            'application/x-rar'
        ]

        if mime_type not in SUPPORTED_TYPES:
            logger.warning(f"Unsupported file type '{mime_type}' for file '{file_name}'. Skipping parsing.")
            return f"НЕПОДДЕРЖИВАЕМЫЙ ТИП ФАЙЛА: {mime_type}"

        if mime_type in ['text/plain', 'text/csv']:
            text = file_content.decode('utf-8', errors='ignore')
            logger.info(f"Successfully decoded text/csv file '{file_name}' with {len(text)} characters.")
            return text

        elif mime_type == 'application/pdf':
            reader = PdfReader(io.BytesIO(file_content))
            text = "".join(page.extract_text() or "" for page in reader.pages)
            logger.info(f"Successfully extracted {len(text)} characters from PDF '{file_name}'.")
            return text

        elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
            doc = Document(io.BytesIO(file_content))
            text = "\n".join([para.text for para in doc.paragraphs])
            logger.info(f"Successfully extracted {len(text)} characters from DOCX '{file_name}'.")
            return text

        elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            workbook = load_workbook(filename=io.BytesIO(file_content))
            text = ""
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    row_text = " ".join([str(cell.value) for cell in row if cell.value is not None])
                    text += row_text + "\n"
            logger.info(f"Successfully extracted {len(text)} characters from XLSX '{file_name}'.")
            return text

        elif mime_type == 'application/zip':
            logger.info(f"[{file_name}] Entering ZIP handler.")
            all_texts = []
            try:
                logger.info(f"[{file_name}] Attempting to open ZIP file in memory.")
                with zipfile.ZipFile(io.BytesIO(file_content)) as zf:
                    logger.info(f"[{file_name}] ZIP file opened successfully. Found {len(zf.infolist())} items inside.")
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        logger.info(f"[{file_name}] -> Processing inner file: '{info.filename}'")
                        inner_file_content = zf.read(info.filename)
                        inner_mime_type = magic.from_buffer(inner_file_content, mime=True)
                        logger.info(f"[{file_name}] -> Inner file '{info.filename}' has MIME type: {inner_mime_type}")
                        inner_text = parse_document(info.filename, inner_file_content, inner_mime_type)
                        logger.info(f"[{file_name}] -> Finished recursive parsing for '{info.filename}'.")
                        if not inner_text.startswith("НЕПОДДЕРЖИВАЕМЫЙ ТИП ФАЙЛА"):
                            all_texts.append(inner_text)
                if all_texts:
                    return "\n\n--- [Content from " + file_name + "] ---\n\n".join(all_texts)
                return f"Архив '{file_name}' не содержит поддерживаемых для анализа файлов."
            except zipfile.BadZipFile as e:
                logger.error(f"Could not process ZIP file {file_name}: {e}")
                return f"Ошибка обработки ZIP архива: {file_name}"

        elif mime_type in ['application/x-rar-compressed', 'application/x-rar']:
            logger.info(f"[{file_name}] Entering RAR handler.")
            all_texts = []
            try:
                logger.info(f"[{file_name}] Attempting to open RAR file in memory.")
                with rarfile.RarFile(io.BytesIO(file_content)) as rf:
                    logger.info(f"[{file_name}] RAR file opened successfully. Found {len(rf.infolist())} items inside.")
                    for info in rf.infolist():
                        if info.is_dir():
                            continue
                        logger.info(f"[{file_name}] -> Processing inner file: '{info.filename}'")
                        inner_file_content = rf.read(info.filename)
                        inner_mime_type = magic.from_buffer(inner_file_content, mime=True)
                        logger.info(f"[{file_name}] -> Inner file '{info.filename}' has MIME type: {inner_mime_type}")
                        inner_text = parse_document(info.filename, inner_file_content, inner_mime_type)
                        logger.info(f"[{file_name}] -> Finished recursive parsing for '{info.filename}'.")
                        if not inner_text.startswith("НЕПОДДЕРЖИВАЕМЫЙ ТИП ФАЙЛА"):
                            all_texts.append(inner_text)
                if all_texts:
                    return "\n\n--- [Content from " + file_name + "] ---\n\n".join(all_texts)
                return f"Архив '{file_name}' не содержит поддерживаемых для анализа файлов."
            except rarfile.Error as e:
                logger.error(f"Could not process RAR file {file_name}: {e}")
                return f"Ошибка обработки RAR архива: {file_name}"

    except Exception as e:
        logger.error(f"Failed to extract text from file '{file_name}' with MIME type {mime_type}. Error: {e}", exc_info=True)
        return f"[Error processing file: {e}]"
    finally:
        logger.info(f"--- PARSER END: Finished parsing '{file_name}' ---")

    return ""
